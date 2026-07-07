"""Export service for generating Excel and PDF reports."""
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.models.awb_models import Document, Shipment, Contact
from app.services.awb_parser import AWBParser


class ExportService:
    """Service for exporting data to Excel and PDF formats."""
    
    # Styling constants
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def export_documents_to_excel(
        self,
        documents: List[Document],
        title: str = "AWB Documents Export"
    ) -> io.BytesIO:
        """Export documents to Excel format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Documents"
        
        # Headers
        headers = [
            "ID", "Document Number", "Master Doc Number", "Reference",
            "Status", "Type", "Shipper", "Consignee",
            "Origin", "Destination", "Route",
            "Document Date", "Created Date"
        ]
        
        self._write_excel_header(ws, headers, title)
        
        # Data rows
        for row_num, doc in enumerate(documents, start=4):
            ws.cell(row=row_num, column=1, value=doc.id)
            ws.cell(row=row_num, column=2, value=doc.document_number)
            ws.cell(row=row_num, column=3, value=doc.master_document_number)
            ws.cell(row=row_num, column=4, value=doc.reference_number)
            ws.cell(row=row_num, column=5, value=self._get_status_name(doc.status))
            ws.cell(row=row_num, column=6, value=self._get_type_name(doc.document_type))
            ws.cell(row=row_num, column=7, value=doc.shipper)
            ws.cell(row=row_num, column=8, value=doc.consignee)
            ws.cell(row=row_num, column=9, value=doc.origin)
            ws.cell(row=row_num, column=10, value=doc.destination)
            ws.cell(row=row_num, column=11, value=doc.route)
            ws.cell(row=row_num, column=12, value=self._format_date(doc.document_date))
            ws.cell(row=row_num, column=13, value=self._format_date(doc.date_created))
            
            # Apply border to data cells
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).border = self.BORDER
        
        # Adjust column widths
        self._auto_adjust_columns(ws)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def export_shipments_to_excel(
        self,
        shipments: List[Shipment],
        title: str = "Shipments Export"
    ) -> io.BytesIO:
        """Export shipments to Excel format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Shipments"
        
        # Headers
        headers = [
            "ID", "Master Number", "House Number", "Reference",
            "Shipper", "Consignee", "Agent",
            "Origin", "Destination", "Shipment Date",
            "Status", "Import/Export"
        ]
        
        self._write_excel_header(ws, headers, title)
        
        # Data rows
        for row_num, ship in enumerate(shipments, start=4):
            ws.cell(row=row_num, column=1, value=ship.id)
            ws.cell(row=row_num, column=2, value=ship.master_number)
            ws.cell(row=row_num, column=3, value=ship.house_number)
            ws.cell(row=row_num, column=4, value=ship.reference_number)
            ws.cell(row=row_num, column=5, value=ship.shipper)
            ws.cell(row=row_num, column=6, value=ship.consignee)
            ws.cell(row=row_num, column=7, value=ship.agent)
            ws.cell(row=row_num, column=8, value=ship.origin)
            ws.cell(row=row_num, column=9, value=ship.destination)
            ws.cell(row=row_num, column=10, value=str(ship.shipment_date) if ship.shipment_date else "")
            ws.cell(row=row_num, column=11, value=ship.event_status)
            ws.cell(row=row_num, column=12, value=self._get_import_export_name(ship.import_export))
            
            # Apply border
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).border = self.BORDER
        
        self._auto_adjust_columns(ws)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def export_contacts_to_excel(
        self,
        contacts: List[Contact],
        title: str = "Contacts Export"
    ) -> io.BytesIO:
        """Export contacts to Excel format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        
        headers = ["ID", "Display Name", "Account Number", "Contact Type", "Station ID"]
        
        self._write_excel_header(ws, headers, title)
        
        for row_num, contact in enumerate(contacts, start=4):
            ws.cell(row=row_num, column=1, value=contact.id)
            ws.cell(row=row_num, column=2, value=contact.display_name)
            ws.cell(row=row_num, column=3, value=contact.account_number)
            ws.cell(row=row_num, column=4, value=self._get_contact_type_name(contact.contact_type))
            ws.cell(row=row_num, column=5, value=contact.station_id)
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).border = self.BORDER
        
        self._auto_adjust_columns(ws)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def export_documents_to_pdf(
        self,
        documents: List[Document],
        title: str = "AWB Documents Report"
    ) -> io.BytesIO:
        """Export documents to PDF format."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1  # Center
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 20))
        
        # Table data
        headers = ["AWB Number", "Shipper", "Consignee", "Origin", "Dest", "Date", "Status"]
        data = [headers]
        
        for doc_item in documents:
            data.append([
                doc_item.document_number or "-",
                (doc_item.shipper or "-")[:30],
                (doc_item.consignee or "-")[:30],
                doc_item.origin or "-",
                doc_item.destination or "-",
                self._format_date(doc_item.document_date),
                self._get_status_name(doc_item.status)
            ])
        
        # Create table
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(
            f"Total Records: {len(documents)}",
            styles['Normal']
        ))
        
        doc.build(elements)
        buffer.seek(0)
        
        return buffer
    
    def export_statistics_to_pdf(
        self,
        stats: Dict[str, Any],
        title: str = "AWB Statistics Report"
    ) -> io.BytesIO:
        """Export statistics to PDF format."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=50,
            leftMargin=50,
            topMargin=50,
            bottomMargin=50
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(
            f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 30))
        
        # Summary section
        elements.append(Paragraph("Summary", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        summary_data = [
            ["Total Documents", str(stats.get('total_documents', 0))],
            ["Total Shipments", str(stats.get('total_shipments', 0))],
            ["Total Contacts", str(stats.get('total_contacts', 0))],
            ["Documents Today", str(stats.get('documents_today', 0))],
            ["Documents This Week", str(stats.get('documents_this_week', 0))],
            ["Documents This Month", str(stats.get('documents_this_month', 0))],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8E8E8')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(summary_table)
        
        doc.build(elements)
        buffer.seek(0)
        
        return buffer
    
    def _write_excel_header(self, ws, headers: List[str], title: str):
        """Write Excel header row with styling."""
        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Date row
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        date_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        date_cell.alignment = Alignment(horizontal='center')
        
        # Header row
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _format_date(dt) -> str:
        """Format datetime for export."""
        if not dt:
            return "-"
        if isinstance(dt, datetime):
            return dt.strftime('%Y-%m-%d')
        return str(dt)
    
    @staticmethod
    def _get_status_name(status: int) -> str:
        """Get document status name."""
        statuses = {0: "Draft", 1: "Active", 2: "Completed", 3: "Cancelled", 4: "Archived"}
        return statuses.get(status, f"Unknown ({status})")
    
    @staticmethod
    def _get_type_name(doc_type: Optional[int]) -> str:
        """Get document type name."""
        if doc_type is None:
            return "-"
        types = {1: "AWB", 2: "HAWB", 3: "MAWB", 4: "Invoice", 5: "Packing List"}
        return types.get(doc_type, f"Type {doc_type}")
    
    @staticmethod
    def _get_import_export_name(ie_type: Optional[int]) -> str:
        """Get import/export type name."""
        if ie_type is None:
            return "-"
        types = {1: "Import", 2: "Export", 3: "Transit"}
        return types.get(ie_type, f"Type {ie_type}")
    
    @staticmethod
    def _get_contact_type_name(contact_type: int) -> str:
        """Get contact type name."""
        types = {1: "Shipper", 2: "Consignee", 3: "Agent", 4: "Notify", 5: "Forwarder"}
        return types.get(contact_type, f"Type {contact_type}")
    
    def export_detailed_awb_report_excel(
        self,
        documents: List[Document],
        title: str = "Rapport AWB Détaillé"
    ) -> io.BytesIO:
        """
        Export AWB documents with full rate description details to Excel.
        Includes pieces, weights, charges, routing, etc.
        """
        wb = Workbook()
        
        # ========== Feuille 1: Récapitulatif ==========
        ws_summary = wb.active
        ws_summary.title = "Récapitulatif"
        
        headers_summary = [
            "N° AWB", "Référence", "Date", "Expéditeur", "Destinataire",
            "Origine", "Destination", "Pièces", "Poids Brut (kg)", 
            "Poids Taxable (kg)", "Devise", "Total Fret", 
            "Autres Frais", "Total Prépayé", "Statut"
        ]
        
        self._write_excel_header(ws_summary, headers_summary, title)
        
        # Parse all documents and collect data
        parsed_docs = []
        total_pieces = 0
        total_gross_weight = 0
        total_chargeable_weight = 0
        total_freight = 0
        total_other_charges = 0
        total_prepaid = 0
        
        for row_num, doc in enumerate(documents, start=4):
            awb_details = AWBParser.parse(doc.document_data) if doc.document_data else None
            parsed_docs.append((doc, awb_details))
            
            pieces = 0
            gross_weight = 0
            chargeable_weight = 0
            currency = ""
            items_total = 0
            other_charges_sum = 0
            prepaid_total = 0
            
            if awb_details:
                pieces = awb_details.total_pieces
                gross_weight = awb_details.total_weight
                chargeable_weight = sum(item.chargeable_weight for item in awb_details.items)
                currency = awb_details.currency
                items_total = awb_details.items_total
                other_charges_sum = sum(charge.amount for charge in awb_details.other_charges)
                prepaid_total = awb_details.charges_summary.total_prepaid
                
                total_pieces += pieces
                total_gross_weight += gross_weight
                total_chargeable_weight += chargeable_weight
                total_freight += items_total
                total_other_charges += other_charges_sum
                total_prepaid += prepaid_total
            
            # Get clean origin/destination codes
            origin_code = doc.origin or ""
            dest_code = doc.destination or ""
            if awb_details:
                # Prefer the airport code from parsed data
                if awb_details.airport_departure_code:
                    origin_code = awb_details.airport_departure_code[:3].upper()
                elif origin_code and len(origin_code) >= 3:
                    origin_code = origin_code[:3].upper()
                # Get destination from route if available
                if awb_details.route.to:
                    dest_code = awb_details.route.to[-1][:3].upper() if awb_details.route.to[-1] else dest_code
                elif dest_code and len(dest_code) >= 3:
                    dest_code = dest_code[:3].upper()
            
            ws_summary.cell(row=row_num, column=1, value=doc.document_number)
            ws_summary.cell(row=row_num, column=2, value=doc.reference_number)
            ws_summary.cell(row=row_num, column=3, value=self._format_date(doc.document_date))
            ws_summary.cell(row=row_num, column=4, value=doc.shipper)
            ws_summary.cell(row=row_num, column=5, value=doc.consignee)
            ws_summary.cell(row=row_num, column=6, value=origin_code)
            ws_summary.cell(row=row_num, column=7, value=dest_code)
            ws_summary.cell(row=row_num, column=8, value=pieces)
            ws_summary.cell(row=row_num, column=9, value=gross_weight)
            ws_summary.cell(row=row_num, column=10, value=chargeable_weight)
            ws_summary.cell(row=row_num, column=11, value=currency)
            ws_summary.cell(row=row_num, column=12, value=items_total)
            ws_summary.cell(row=row_num, column=13, value=other_charges_sum)
            ws_summary.cell(row=row_num, column=14, value=prepaid_total)
            ws_summary.cell(row=row_num, column=15, value=self._get_status_name(doc.status))
            
            for col in range(1, len(headers_summary) + 1):
                ws_summary.cell(row=row_num, column=col).border = self.BORDER
        
        # Add totals row
        total_row = len(documents) + 4
        ws_summary.cell(row=total_row, column=1, value="TOTAUX")
        ws_summary.cell(row=total_row, column=1).font = Font(bold=True)
        ws_summary.cell(row=total_row, column=8, value=total_pieces).font = Font(bold=True)
        ws_summary.cell(row=total_row, column=9, value=total_gross_weight).font = Font(bold=True)
        ws_summary.cell(row=total_row, column=10, value=total_chargeable_weight).font = Font(bold=True)
        ws_summary.cell(row=total_row, column=12, value=total_freight).font = Font(bold=True)
        ws_summary.cell(row=total_row, column=13, value=total_other_charges).font = Font(bold=True)
        ws_summary.cell(row=total_row, column=14, value=total_prepaid).font = Font(bold=True)
        
        for col in range(1, len(headers_summary) + 1):
            cell = ws_summary.cell(row=total_row, column=col)
            cell.border = self.BORDER
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        self._auto_adjust_columns(ws_summary)
        
        # ========== Feuille 2: Détail des lignes de tarification ==========
        ws_items = wb.create_sheet("Lignes de tarification")
        
        headers_items = [
            "N° AWB", "Date", "Expéditeur", "Pièces", "Poids Brut",
            "Poids Taxable", "Tarif Unitaire", "Total Ligne", 
            "Nature des marchandises", "Devise"
        ]
        
        self._write_excel_header(ws_items, headers_items, "Détail des Lignes de Tarification")
        
        item_row = 4
        for doc, awb_details in parsed_docs:
            if awb_details and awb_details.items:
                for item in awb_details.items:
                    ws_items.cell(row=item_row, column=1, value=doc.document_number)
                    ws_items.cell(row=item_row, column=2, value=self._format_date(doc.document_date))
                    ws_items.cell(row=item_row, column=3, value=doc.shipper)
                    ws_items.cell(row=item_row, column=4, value=item.pieces)
                    ws_items.cell(row=item_row, column=5, value=item.gross_weight)
                    ws_items.cell(row=item_row, column=6, value=item.chargeable_weight)
                    ws_items.cell(row=item_row, column=7, value=item.rate_charge)
                    ws_items.cell(row=item_row, column=8, value=item.total)
                    ws_items.cell(row=item_row, column=9, value=item.nature[:100] if item.nature else "")
                    ws_items.cell(row=item_row, column=10, value=awb_details.currency)
                    
                    for col in range(1, len(headers_items) + 1):
                        ws_items.cell(row=item_row, column=col).border = self.BORDER
                    
                    item_row += 1
        
        self._auto_adjust_columns(ws_items)
        
        # ========== Feuille 3: Autres frais ==========
        ws_charges = wb.create_sheet("Autres Frais")
        
        headers_charges = [
            "N° AWB", "Date", "Code Frais", "Description", "Montant", 
            "Dû à", "Devise"
        ]
        
        self._write_excel_header(ws_charges, headers_charges, "Autres Frais et Taxes")
        
        charge_code_names = {
            "PS": "Frais de sécurité",
            "AWC": "Air Waybill Charge",
            "FC": "Fuel Surcharge",
            "SC": "Security Charge",
            "MY": "Miscellaneous Charges",
            "RA": "Dangerous Goods",
            "SR": "Screen Charge",
            "AW": "AWB Fee",
        }
        
        charge_row = 4
        for doc, awb_details in parsed_docs:
            if awb_details and awb_details.other_charges:
                for charge in awb_details.other_charges:
                    ws_charges.cell(row=charge_row, column=1, value=doc.document_number)
                    ws_charges.cell(row=charge_row, column=2, value=self._format_date(doc.document_date))
                    ws_charges.cell(row=charge_row, column=3, value=charge.code)
                    ws_charges.cell(row=charge_row, column=4, value=charge_code_names.get(charge.code, charge.code))
                    ws_charges.cell(row=charge_row, column=5, value=charge.amount)
                    ws_charges.cell(row=charge_row, column=6, value=charge.due)
                    ws_charges.cell(row=charge_row, column=7, value=awb_details.currency)
                    
                    for col in range(1, len(headers_charges) + 1):
                        ws_charges.cell(row=charge_row, column=col).border = self.BORDER
                    
                    charge_row += 1
        
        self._auto_adjust_columns(ws_charges)
        
        # ========== Feuille 4: Routing/Vols ==========
        ws_routing = wb.create_sheet("Routing")
        
        headers_routing = [
            "N° AWB", "Date", "Origine", "Destination", "Escales", 
            "Compagnies", "Vols", "Manutention"
        ]
        
        self._write_excel_header(ws_routing, headers_routing, "Informations de Routing")
        
        routing_row = 4
        for doc, awb_details in parsed_docs:
            via = ""
            carriers = ""
            flights = ""
            handling = ""
            origin_code = doc.origin or ""
            dest_code = doc.destination or ""
            
            if awb_details:
                if awb_details.route.to:
                    via = " → ".join(filter(None, awb_details.route.to))
                if awb_details.route.by:
                    carriers = ", ".join(filter(None, awb_details.route.by))
                if awb_details.route.flights:
                    flights = ", ".join(filter(None, awb_details.route.flights))
                handling = awb_details.handling_information
                
                # Clean origin/destination codes
                if awb_details.airport_departure_code:
                    origin_code = awb_details.airport_departure_code[:3].upper()
                elif origin_code and len(origin_code) >= 3:
                    origin_code = origin_code[:3].upper()
                if awb_details.route.to:
                    dest_code = awb_details.route.to[-1][:3].upper() if awb_details.route.to[-1] else dest_code
                elif dest_code and len(dest_code) >= 3:
                    dest_code = dest_code[:3].upper()
            
            ws_routing.cell(row=routing_row, column=1, value=doc.document_number)
            ws_routing.cell(row=routing_row, column=2, value=self._format_date(doc.document_date))
            ws_routing.cell(row=routing_row, column=3, value=origin_code)
            ws_routing.cell(row=routing_row, column=4, value=dest_code)
            ws_routing.cell(row=routing_row, column=5, value=via)
            ws_routing.cell(row=routing_row, column=6, value=carriers)
            ws_routing.cell(row=routing_row, column=7, value=flights)
            ws_routing.cell(row=routing_row, column=8, value=handling)
            
            for col in range(1, len(headers_routing) + 1):
                ws_routing.cell(row=routing_row, column=col).border = self.BORDER
            
            routing_row += 1
        
        self._auto_adjust_columns(ws_routing)
        
        # ========== Feuille 5: Statistiques ==========
        ws_stats = wb.create_sheet("Statistiques")
        
        # Title
        ws_stats.merge_cells('A1:D1')
        ws_stats.cell(row=1, column=1, value="Statistiques du Rapport")
        ws_stats.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        ws_stats.cell(row=3, column=1, value="Métrique")
        ws_stats.cell(row=3, column=2, value="Valeur")
        ws_stats.cell(row=3, column=1).font = Font(bold=True)
        ws_stats.cell(row=3, column=2).font = Font(bold=True)
        
        stats_data = [
            ("Nombre de documents", len(documents)),
            ("Total pièces", total_pieces),
            ("Poids brut total (kg)", round(total_gross_weight, 2)),
            ("Poids taxable total (kg)", round(total_chargeable_weight, 2)),
            ("Total fret", round(total_freight, 2)),
            ("Total autres frais", round(total_other_charges, 2)),
            ("Total prépayé", round(total_prepaid, 2)),
            ("Moyenne pièces/AWB", round(total_pieces / len(documents), 1) if documents else 0),
            ("Moyenne poids/AWB (kg)", round(total_gross_weight / len(documents), 1) if documents else 0),
            ("Moyenne montant/AWB", round(total_prepaid / len(documents), 2) if documents else 0),
        ]
        
        for idx, (metric, value) in enumerate(stats_data, start=4):
            ws_stats.cell(row=idx, column=1, value=metric)
            ws_stats.cell(row=idx, column=2, value=value)
            ws_stats.cell(row=idx, column=1).border = self.BORDER
            ws_stats.cell(row=idx, column=2).border = self.BORDER
        
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 20
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def export_detailed_awb_report_pdf(
        self,
        documents: List[Document],
        stats: Dict[str, Any],
        title: str = "Rapport d'activite",
        total_count: Optional[int] = None,
        period_start=None,
        period_end=None,
    ) -> io.BytesIO:
        """Rapport d'activite executif (portrait A4) pour direction & comptabilite."""
        from collections import defaultdict
        from reportlab.platypus import PageBreak, HRFlowable
        from app.services.statistics_service import StatisticsService

        BRAND = colors.HexColor('#0c6d61')
        BRAND_DK = colors.HexColor('#0a574e')
        LIGHT = colors.HexColor('#e8f2f0')
        ROW_ALT = colors.HexColor('#f5f8f7')
        GREY = colors.HexColor('#5b6b7f')
        INK = colors.HexColor('#13273b')
        LINE = colors.HexColor('#dbe4e1')
        WHITE = colors.white

        STATUS_FR = {0: "Brouillon", 1: "Actif", 2: "Complété", 3: "Annulé", 4: "Archivé"}

        def fi(v):
            try:
                return f"{int(round(float(v))):,}".replace(",", " ")
            except Exception:
                return "0"

        def fm(v):
            try:
                return f"{float(v):,.2f}".replace(",", "§").replace(".", ",").replace("§", " ")
            except Exception:
                return "0,00"

        def fp(v):
            try:
                return f"{float(v):.1f}".replace(".", ",")
            except Exception:
                return "0,0"

        # ---------- Agregation ----------
        recs = []
        for d in documents:
            awb = AWBParser.parse(d.document_data) if d.document_data else None
            num = (d.document_number or '').strip()
            prefix = num[:3] if (len(num) >= 3 and num[:3].isdigit()) else None
            pieces = awb.total_pieces if awb else 0
            weight = float(awb.total_weight) if awb else 0.0
            prepaid = float(awb.charges_summary.total_prepaid) if awb else 0.0
            currency = (awb.currency if (awb and awb.currency) else '')
            dest = (d.destination or '').strip()
            if awb and getattr(awb, 'route', None) and awb.route.to:
                dest = (awb.route.to[-1] or dest)
            dest = (dest[:3].upper() if dest else '-')
            origin = (d.origin or '')
            if awb and getattr(awb, 'airport_departure_code', None):
                origin = awb.airport_departure_code or origin
            origin = (origin[:3].upper() if origin else '-')
            shipper = ((d.shipper or '').strip() or '-')
            recs.append(dict(num=(num or '-'), prefix=prefix, pieces=pieces, weight=weight,
                             prepaid=prepaid, currency=currency, dest=dest, origin=origin,
                             shipper=shipper, consignee=(d.consignee or '-'), status=d.status))

        tot_pieces = sum(r['pieces'] for r in recs)
        tot_weight = sum(r['weight'] for r in recs)
        tot_revenue = sum(r['prepaid'] for r in recs)
        currencies = {r['currency'] for r in recs if r['currency']}
        main_cur = (stats.get('main_currency') or (sorted(currencies)[0] if currencies else 'USD'))
        n_docs = total_count if total_count is not None else len(recs)
        base = len(recs) or 1

        by_air = defaultdict(lambda: {'count': 0, 'rev': 0.0})
        by_dest = defaultdict(lambda: {'count': 0, 'rev': 0.0})
        by_client = defaultdict(lambda: {'count': 0, 'rev': 0.0})
        by_status = defaultdict(int)
        for r in recs:
            if r['prefix']:
                by_air[r['prefix']]['count'] += 1
                by_air[r['prefix']]['rev'] += r['prepaid']
            by_dest[r['dest']]['count'] += 1
            by_dest[r['dest']]['rev'] += r['prepaid']
            by_client[r['shipper']]['count'] += 1
            by_client[r['shipper']]['rev'] += r['prepaid']
            by_status[r['status']] += 1

        # ---------- Document ----------
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=40, leftMargin=40, topMargin=42, bottomMargin=50,
                                title=title, author="Elite Cargo")
        avail = doc.width

        styles = getSampleStyleSheet()
        st_title = ParagraphStyle('t', parent=styles['Normal'], fontSize=17, leading=20, textColor=INK)
        h_sec = ParagraphStyle('sec', parent=styles['Heading2'], fontSize=12, leading=15,
                               textColor=BRAND_DK, spaceAfter=2, spaceBefore=2, fontName='Helvetica-Bold')
        p_small = ParagraphStyle('small', parent=styles['Normal'], fontSize=8, textColor=GREY, leading=11)
        p_cell = ParagraphStyle('pcell', parent=styles['Normal'], fontSize=8.5, textColor=INK, leading=11)

        el = []

        # En-tete bandeau
        header_left = Paragraph(
            '<font size=15 color="#ffffff"><b>ELITE CARGO</b></font><br/>'
            '<font size=8 color="#d5ece7">Gestion de fret aérien · AWB</font>', styles['Normal'])
        header_right = Paragraph(
            '<para align="right"><font size=8 color="#d5ece7">Édité le</font><br/>'
            f'<font size=10 color="#ffffff"><b>{datetime.now().strftime("%d/%m/%Y à %Hh%M")}</b></font></para>',
            styles['Normal'])
        band = Table([[header_left, header_right]], colWidths=[avail * 0.60, avail * 0.40])
        band.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BRAND),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 14),
            ('RIGHTPADDING', (0, 0), (-1, -1), 14),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        el.append(band)
        el.append(Spacer(1, 14))

        if period_start and period_end:
            per = f"Période : du {period_start.strftime('%d/%m/%Y')} au {period_end.strftime('%d/%m/%Y')}"
        elif period_start:
            per = f"Période : depuis le {period_start.strftime('%d/%m/%Y')}"
        elif period_end:
            per = f"Période : jusqu'au {period_end.strftime('%d/%m/%Y')}"
        else:
            per = "Période : toutes les données (cumul)"
        el.append(Paragraph(f'<b>{title}</b>', st_title))
        el.append(Paragraph(per, p_small))
        el.append(Spacer(1, 14))

        # Synthese KPI
        def delta_txt(pct):
            if pct is None:
                return '<font size=7 color="#5b6b7f">n/c vs préc.</font>'
            col = '#059669' if pct >= 0 else '#dc2626'
            sign = '+' if pct >= 0 else ''
            return f'<font size=7 color="{col}">{sign}{fp(pct)} % vs préc.</font>'

        kpi_lab = ParagraphStyle('kl', parent=styles['Normal'], fontSize=7, textColor=GREY, leading=9, spaceAfter=5)
        kpi_val = ParagraphStyle('kv', parent=styles['Normal'], fontSize=17, textColor=INK, leading=20, fontName='Helvetica-Bold', spaceAfter=4)
        kpi_del = ParagraphStyle('kd', parent=styles['Normal'], fontSize=7, leading=9)

        def kpi_cell(label, value, pct):
            return [Paragraph(label, kpi_lab), Paragraph(value, kpi_val), Paragraph(delta_txt(pct), kpi_del)]

        kpi = [[
            kpi_cell("CHIFFRE D'AFFAIRES", f"{fi(tot_revenue)} {main_cur}", stats.get('revenue_change_pct')),
            kpi_cell("LTA (DOCUMENTS)", fi(n_docs), stats.get('documents_change_pct')),
            kpi_cell("POIDS TOTAL", f"{fm(tot_weight / 1000)} t", stats.get('weight_change_pct')),
            kpi_cell("PIÈCES", fi(tot_pieces), stats.get('pieces_change_pct')),
        ]]
        kt = Table(kpi, colWidths=[avail / 4.0] * 4)
        kt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), LIGHT),
            ('BOX', (0, 0), (-1, -1), 0.5, LINE),
            ('INNERGRID', (0, 0), (-1, -1), 1.5, WHITE),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 11),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        el.append(kt)
        el.append(Spacer(1, 6))
        el.append(Paragraph(
            f"CA = total prépayé porté sur les LTA de la période (devise principale {main_cur}). "
            "Évolution comparée à la période précédente de même durée.", p_small))
        el.append(Spacer(1, 16))

        def section(txt):
            el.append(Paragraph(txt, h_sec))
            el.append(HRFlowable(width='100%', thickness=1, color=BRAND, spaceBefore=2, spaceAfter=8))

        def styled_table(header, rows, widths, right_from=1):
            t = Table([header] + rows, colWidths=widths, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BRAND),
                ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8.5),
                ('TEXTCOLOR', (0, 1), (-1, -1), INK),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, ROW_ALT]),
                ('LINEBELOW', (0, 0), (-1, -1), 0.4, LINE),
                ('ALIGN', (right_from, 0), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 7),
                ('RIGHTPADDING', (0, 0), (-1, -1), 7),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            return t

        # Compagnies
        section("Répartition par compagnie aérienne")
        air_sorted = sorted(by_air.items(), key=lambda kv: kv[1]['rev'], reverse=True)[:8]
        rows = [[Paragraph(f"{StatisticsService.AIRLINE_PREFIXES.get(pfx, 'Compagnie (' + pfx + ')')} ({pfx})", p_cell),
                 fi(v['count']), f"{fp(v['count'] / base * 100)} %", f"{fi(v['rev'])} {main_cur}"]
                for pfx, v in air_sorted] or [[Paragraph("-", p_cell), "0", "0,0 %", f"0 {main_cur}"]]
        el.append(styled_table(["Compagnie", "LTA", "Part", f"CA ({main_cur})"], rows,
                               [avail * 0.46, avail * 0.13, avail * 0.16, avail * 0.25]))
        el.append(Spacer(1, 14))

        # Destinations
        section("Principales destinations")
        dest_sorted = sorted(by_dest.items(), key=lambda kv: kv[1]['count'], reverse=True)[:8]
        rows = [[Paragraph(dst, p_cell), fi(v['count']), f"{fp(v['count'] / base * 100)} %", f"{fi(v['rev'])} {main_cur}"]
                for dst, v in dest_sorted] or [[Paragraph("-", p_cell), "0", "0,0 %", f"0 {main_cur}"]]
        el.append(styled_table(["Destination", "LTA", "Part", f"CA ({main_cur})"], rows,
                               [avail * 0.46, avail * 0.13, avail * 0.16, avail * 0.25]))
        el.append(Spacer(1, 14))

        # Clients
        section("Principaux clients (expéditeurs)")
        cl_sorted = sorted(by_client.items(), key=lambda kv: kv[1]['rev'], reverse=True)[:8]
        rows = [[Paragraph(name[:42], p_cell), fi(v['count']), f"{fi(v['rev'])} {main_cur}"]
                for name, v in cl_sorted] or [[Paragraph("-", p_cell), "0", f"0 {main_cur}"]]
        el.append(styled_table(["Client", "LTA", f"CA ({main_cur})"], rows,
                               [avail * 0.60, avail * 0.16, avail * 0.24]))
        el.append(Spacer(1, 14))


        # Annexe detail
        el.append(PageBreak())
        section("Annexe — Détail des LTA")
        shown = min(len(recs), 60)
        note = f"{shown} LTA affichées" + (f" sur {n_docs}." if n_docs > 60 else ".")
        el.append(Paragraph(note, p_small))
        el.append(Spacer(1, 6))
        det_rows = []
        for r in recs[:60]:
            det_rows.append([
                r['num'],
                Paragraph((r['shipper'] or '-')[:26], p_cell),
                Paragraph((r['consignee'] or '-')[:26], p_cell),
                f"{r['origin']} > {r['dest']}",
                fi(r['pieces']), fm(r['weight']), fi(r['prepaid'])])
        if not det_rows:
            det_rows = [["-", Paragraph("-", p_cell), Paragraph("-", p_cell), "-", "0", "0,00", "0"]]
        det = Table([["N° AWB", "Expéditeur", "Destinataire", "Route", "Pièces", "Poids", f"CA {main_cur}"]] + det_rows,
                    repeatRows=1,
                    colWidths=[avail * 0.16, avail * 0.21, avail * 0.21, avail * 0.15, avail * 0.08, avail * 0.10, avail * 0.09])
        det.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), BRAND),
            ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('TEXTCOLOR', (0, 1), (-1, -1), INK),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, ROW_ALT]),
            ('LINEBELOW', (0, 0), (-1, -1), 0.4, LINE),
            ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        el.append(det)

        def _footer(canvas, doc_):
            canvas.saveState()
            canvas.setStrokeColor(LINE)
            canvas.setLineWidth(0.5)
            canvas.line(doc.leftMargin, 36, A4[0] - doc.rightMargin, 36)
            canvas.setFont('Helvetica', 7.5)
            canvas.setFillColor(GREY)
            canvas.drawString(doc.leftMargin, 26, "Elite Cargo · Rapport d'activité · Document confidentiel")
            canvas.drawRightString(A4[0] - doc.rightMargin, 26, f"Page {doc_.page}")
            canvas.restoreState()

        doc.build(el, onFirstPage=_footer, onLaterPages=_footer)
        buffer.seek(0)
        return buffer
